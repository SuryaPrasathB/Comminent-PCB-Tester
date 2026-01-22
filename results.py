# results.py
# Results tab – view & export test results
# Starts EMPTY, loads data ONLY on fetch

import os
import mysql.connector
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QMessageBox,
    QTableWidgetItem, QFileDialog
)
from PySide6.QtUiTools import QUiLoader
from PySide6.QtCore import QFile, QIODevice, QDateTime, QTime, QDate

from config import DB_CONFIG
from logs import logger
from PySide6.QtWidgets import QAbstractItemView

import openpyxl


class ResultsController:
    def __init__(self, parent_tab: QWidget):
        print("[RESULTS] Initializing ResultsController")
        logger.info("Initializing ResultsController")

        self.parent_tab = parent_tab
        self.ui = None
        self.table = None

        self.load_ui()
        self.connect_signals()

    # -------------------------------------------------
    def load_ui(self):
        loader = QUiLoader()
        base_dir = os.path.dirname(os.path.abspath(__file__))
        ui_path = os.path.join(base_dir, "results.ui")

        print("[RESULTS] Loading UI from:", ui_path)
        logger.info(f"Loading results UI from: {ui_path}")

        ui_file = QFile(ui_path)
        if not ui_file.open(QIODevice.ReadOnly):
            print("[RESULTS][ERROR] Unable to open results.ui")
            logger.error("Unable to open results.ui")
            raise RuntimeError(f"Unable to open results.ui: {ui_path}")

        self.ui = loader.load(ui_file, self.parent_tab)
        ui_file.close()

        if not self.ui:
            print("[RESULTS][ERROR] results.ui loaded as NULL")
            logger.error("results.ui loaded as NULL")
            raise RuntimeError("Failed to load results.ui")

        # Remove old layout if any
        old_layout = self.parent_tab.layout()
        if old_layout:
            QWidget().setLayout(old_layout)

        layout = QVBoxLayout(self.parent_tab)
        layout.addWidget(self.ui)
        self.parent_tab.setLayout(layout)

        self.table = self.ui.table_results

        self.table.setEditTriggers(QAbstractItemView.NoEditTriggers)

        # Start EMPTY + HIDDEN
        self.table.setRowCount(0)
        self.table.setVisible(False)

        # ---------------- DEFAULT DATE & TIME ----------------
        today = QDate.currentDate()

        from_time = QTime(0, 0)  # 12:00 AM
        to_time = QTime(23, 59)  # 11:59 PM

        self.ui.dt_from.setDateTime(QDateTime(today, from_time))
        self.ui.dt_to.setDateTime(QDateTime(today, to_time))
        # ----------------------------------------------------

        print("[RESULTS] UI loaded successfully (table hidden)")
        logger.info("Results UI loaded successfully")

    # -------------------------------------------------
    def connect_signals(self):
        print("[RESULTS] Connecting signals")
        logger.info("Connecting ResultsController signals")

        self.ui.btn_fetch_serial.clicked.connect(self.fetch_by_serial)
        self.ui.btn_fetch_date.clicked.connect(self.fetch_by_date)
        self.ui.btn_export.clicked.connect(self.export_results)

    # -------------------------------------------------
    def _db_connect(self):
        print("[RESULTS] Connecting to database")
        logger.info("Connecting to database for results")
        return mysql.connector.connect(**DB_CONFIG)

    # -------------------------------------------------
    def fetch_by_serial(self):
        print("[RESULTS] Fetch by serial clicked")
        logger.info("Fetch by serial clicked")

        serial = self.ui.txt_serial.text().strip()
        print("[RESULTS] Serial entered:", serial)
        logger.info(f"Serial entered: {serial}")

        if not serial:
            QMessageBox.warning(self.ui, "Input", "Enter PCB Serial Number")
            print("[RESULTS] Serial empty → abort")
            logger.warning("Serial empty → fetch aborted")
            return

        query = """
            SELECT project_name, pcb_serial, sn, description,
                   r, y, b, n,
                   expected_v, expected_i,
                   measured_v, measured_i,
                   result, tested_at
            FROM test_results
            WHERE pcb_serial = %s
            ORDER BY tested_at DESC
        """

        self._run_query(query, (serial,))

    # -------------------------------------------------
    def fetch_by_date(self):
        print("[RESULTS] Fetch by date clicked")
        logger.info("Fetch by date clicked")

        from_dt = self.ui.dt_from.dateTime().toPython()
        to_dt   = self.ui.dt_to.dateTime().toPython()

        print("[RESULTS] Date range:", from_dt, "→", to_dt)
        logger.info(f"Date range: {from_dt} → {to_dt}")

        query = """
            SELECT project_name, pcb_serial, sn, description,
                   r, y, b, n,
                   expected_v, expected_i,
                   measured_v, measured_i,
                   result, tested_at
            FROM test_results
            WHERE tested_at BETWEEN %s AND %s
            ORDER BY tested_at DESC
        """

        self._run_query(query, (from_dt, to_dt))

    # -------------------------------------------------
    def _run_query(self, query, params):
        try:
            print("[RESULTS] Running database query")
            logger.info("Running database query")

            conn = self._db_connect()
            cur = conn.cursor()

            print("[RESULTS] Executing SQL")
            logger.info("Executing SQL query")

            cur.execute(query, params)

            rows = cur.fetchall()
            print("[RESULTS] Rows fetched:", len(rows))
            logger.info(f"Rows fetched: {len(rows)}")

            cur.close()
            conn.close()

            if not rows:
                print("[RESULTS] No data found")
                logger.warning("No results found for query")
                QMessageBox.information(
                    self.ui,
                    "Results",
                    "No results found for given criteria"
                )
                self.table.setRowCount(0)
                self.table.setVisible(False)
                return

            self.populate_table(rows)
            self.table.setVisible(True)

            print("[RESULTS] Table populated and shown")
            logger.info("Results table populated and shown")

        except Exception as e:
            print("[RESULTS][ERROR] Database error:", e)
            logger.error(f"Database error: {e}")
            QMessageBox.critical(self.ui, "Database Error", str(e))

    # -------------------------------------------------
    def populate_table(self, rows):
        print("[RESULTS] Populating table")
        logger.info("Populating results table")

        self.table.setRowCount(0)

        for r, row in enumerate(rows):
            self.table.insertRow(r)
            for c, value in enumerate(row):
                self.table.setItem(
                    r, c, QTableWidgetItem(str(value))
                )

        print("[RESULTS] Table rows:", self.table.rowCount())
        logger.info(f"Table rows populated: {self.table.rowCount()}")

    # -------------------------------------------------
    def export_results(self):
        print("[RESULTS] Export button clicked")
        logger.info("Export button clicked")

        if self.table.rowCount() == 0:
            print("[RESULTS] No data to export")
            logger.warning("Export aborted: no data")
            QMessageBox.warning(
                self.ui,
                "Export",
                "No data to export. Fetch results first."
            )
            return

        pdf   = self.ui.chk_pdf.isChecked()
        excel = self.ui.chk_excel.isChecked()

        print("[RESULTS] Export options → PDF:", pdf, "Excel:", excel)
        logger.info(f"Export options → PDF={pdf}, Excel={excel}")

        if not pdf and not excel:
            QMessageBox.warning(self.ui, "Export", "Select PDF or Excel")
            logger.warning("Export aborted: no format selected")
            return

        if excel:
            self.export_excel()
        if pdf:
            self.export_pdf()

    # -------------------------------------------------
    def export_excel(self):
        print("[RESULTS] Exporting to Excel")
        logger.info("Exporting results to Excel")

        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from datetime import datetime
        import pandas as pd
        import os

        # -------------------------------------------------
        # Select file
        # -------------------------------------------------
        path, _ = QFileDialog.getSaveFileName(
            self.ui, "Save Excel", "", "Excel Files (*.xlsx)"
        )
        if not path:
            logger.warning("Excel export cancelled")
            return

        # -------------------------------------------------
        # Handle locked file
        # -------------------------------------------------
        if os.path.exists(path):
            try:
                with open(path, "ab"):
                    pass
            except PermissionError:
                base, ext = os.path.splitext(path)
                path = f"{base}_{datetime.now().strftime('%H%M%S')}{ext}"
                logger.warning("Excel locked → auto-renamed")

        try:
            # -------------------------------------------------
            # READ TABLE HEADERS
            # -------------------------------------------------
            all_headers = [
                self.table.horizontalHeaderItem(i).text().strip()
                for i in range(self.table.columnCount())
            ]

            # -------------------------------------------------
            # COLUMNS TO SKIP (SAME AS PDF)
            # -------------------------------------------------
            SKIP_COLUMNS = {"r", "y", "b", "n"}

            # -------------------------------------------------
            # EXCEL COLUMN WIDTHS
            # -------------------------------------------------
            COLUMN_WIDTHS = {
                "project": 18,
                "pcb serial": 22,
                "sn": 8,
                "description": 35,

                "exp v": 10,
                "exp i": 10,
                "meas v": 10,
                "meas i": 10,

                "result": 10,
                "tested at": 22,
            }

            # -------------------------------------------------
            # NUMERIC COLUMNS (IMPORTANT FIX)
            # -------------------------------------------------
            NUMERIC_COLUMNS = {
                "sn",
                "exp v",
                "exp i",
                "meas v",
                "meas i",
            }

            excel_columns = []  # (table_col_index, header, width)

            for idx, h in enumerate(all_headers):
                hl = h.lower()

                if hl in SKIP_COLUMNS:
                    continue

                if hl not in COLUMN_WIDTHS:
                    QMessageBox.critical(
                        self.ui,
                        "Export Error",
                        f"No Excel width defined for column:\n'{h}'"
                    )
                    logger.error(f"No Excel width defined for '{h}'")
                    return

                excel_columns.append((idx, h, COLUMN_WIDTHS[hl]))

            # -------------------------------------------------
            # COLLECT DATA (WITH TYPE CONVERSION)
            # -------------------------------------------------
            data = []

            for r in range(self.table.rowCount()):
                row_data = []
                for col_idx, header, _ in excel_columns:
                    item = self.table.item(r, col_idx)
                    text = item.text().strip() if item else ""
                    hl = header.lower()

                    if hl in NUMERIC_COLUMNS:
                        try:
                            if "." in text:
                                row_data.append(float(text))
                            else:
                                row_data.append(int(text))
                        except ValueError:
                            row_data.append(text)
                    else:
                        row_data.append(text)

                data.append(row_data)

            headers = [h for _, h, _ in excel_columns]
            df = pd.DataFrame(data, columns=headers)

            # -------------------------------------------------
            # WRITE EXCEL USING OPENPYXL
            # -------------------------------------------------
            with pd.ExcelWriter(path, engine="openpyxl") as writer:
                df.to_excel(
                    writer,
                    index=False,
                    startrow=4,
                    sheet_name="Test Results"
                )

                ws = writer.sheets["Test Results"]

                # -------------------------------------------------
                # TITLE & META
                # -------------------------------------------------
                ws.merge_cells(
                    start_row=1, start_column=1,
                    end_row=1, end_column=len(headers)
                )
                ws.cell(row=1, column=1).value = "PCB TEST RESULTS REPORT"

                ws.cell(row=3, column=1).value = "Date & Time:"
                ws.cell(row=3, column=2).value = datetime.now().strftime(
                    "%d-%m-%Y %H:%M:%S"
                )

                ws.cell(row=4, column=1).value = "Generated By:"
                ws.cell(row=4, column=2).value = "Comminent PCB Tester"

                # -------------------------------------------------
                # SET COLUMN WIDTHS
                # -------------------------------------------------
                for col_num, (_, _, width) in enumerate(excel_columns, start=1):
                    ws.column_dimensions[
                        ws.cell(row=5, column=col_num).column_letter
                    ].width = width

            logger.info(f"Excel exported: {path}")
            QMessageBox.information(
                self.ui, "Export", "Excel exported successfully"
            )

        except PermissionError:
            logger.error("Excel permission denied")
            QMessageBox.critical(
                self.ui,
                "Export Failed",
                "Close the Excel file and try again."
            )

        except Exception as e:
            logger.error(f"Excel export failed: {e}")
            QMessageBox.critical(
                self.ui,
                "Export Failed",
                str(e)
            )

    # -------------------------------------------------
    def export_pdf(self):
        print("[RESULTS] Exporting to PDF")
        logger.info("Exporting results to PDF")

        from PySide6.QtWidgets import QFileDialog, QMessageBox
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.pdfgen import canvas
        from datetime import datetime

        # -------------------------------------------------
        # Select file
        # -------------------------------------------------
        path, _ = QFileDialog.getSaveFileName(
            self.ui, "Save PDF", "", "PDF Files (*.pdf)"
        )
        if not path:
            logger.warning("PDF export cancelled")
            return

        try:
            # =================================================
            # LANDSCAPE PAGE (OVERWRITE MODE)
            # =================================================
            c = canvas.Canvas(path, pagesize=landscape(A4))
            page_width, page_height = landscape(A4)

            # -------------------------------------------------
            # HEADER
            # -------------------------------------------------
            c.setFont("Helvetica-Bold", 18)
            c.drawCentredString(
                page_width / 2, page_height - 40,
                "PCB TEST RESULTS REPORT"
            )

            c.setFont("Helvetica", 10)
            c.drawString(
                40, page_height - 70,
                f"Date & Time : {datetime.now().strftime('%d-%m-%Y %H:%M:%S')}"
            )
            c.drawString(
                40, page_height - 90,
                "Generated By : Comminent PCB Tester"
            )

            # -------------------------------------------------
            # TABLE SETUP
            # -------------------------------------------------
            start_x = 30
            start_y = page_height - 130
            row_height = 20

            all_headers = [
                self.table.horizontalHeaderItem(i).text().strip()
                for i in range(self.table.columnCount())
            ]

            # -------------------------------------------------
            # COLUMNS TO SKIP IN PDF
            # -------------------------------------------------
            SKIP_COLUMNS = {"r", "y", "b", "n"}

            # -------------------------------------------------
            # PDF COLUMN WIDTH DEFINITIONS
            # -------------------------------------------------
            COLUMN_WIDTHS = {
                "project": 90,
                "pcb serial": 110,
                "sn": 40,
                "description": 150,

                "exp v": 45,
                "exp i": 45,
                "meas v": 45,
                "meas i": 45,

                "result": 40,
                "tested at": 110,
            }

            pdf_columns = []  # (table_col_index, header, width)

            for idx, h in enumerate(all_headers):
                hl = h.lower()

                if hl in SKIP_COLUMNS:
                    continue

                if hl not in COLUMN_WIDTHS:
                    QMessageBox.critical(
                        self.ui,
                        "Export Error",
                        f"No PDF width defined for column:\n'{h}'"
                    )
                    logger.error(f"No PDF width defined for '{h}'")
                    return

                pdf_columns.append((idx, h, COLUMN_WIDTHS[hl]))

            # -------------------------------------------------
            # DRAW HEADER ROW
            # -------------------------------------------------
            c.setFont("Helvetica-Bold", 9)
            x = start_x
            y = start_y

            for _, header, width in pdf_columns:
                c.rect(x, y - row_height, width, row_height)
                c.drawString(x + 4, y - 14, header)
                x += width

            y -= row_height
            c.setFont("Helvetica", 9)

            # -------------------------------------------------
            # DRAW DATA ROWS
            # -------------------------------------------------
            for r in range(self.table.rowCount()):
                x = start_x

                for col_idx, _, width in pdf_columns:
                    c.rect(x, y - row_height, width, row_height)

                    item = self.table.item(r, col_idx)
                    text = item.text() if item else ""

                    c.drawString(x + 4, y - 14, text[:80])
                    x += width

                y -= row_height

                if y < 60:
                    c.showPage()
                    c.setFont("Helvetica", 9)
                    y = page_height - 60

            # -------------------------------------------------
            # FOOTER
            # -------------------------------------------------
            c.setFont("Helvetica-Oblique", 8)
            c.drawCentredString(
                page_width / 2, 30,
                "Report generated automatically"
            )

            # -------------------------------------------------
            # SAVE (OVERWRITE HAPPENS HERE)
            # -------------------------------------------------
            c.save()

            logger.info(f"PDF exported (overwritten if existed): {path}")
            QMessageBox.information(
                self.ui, "Export", "PDF exported successfully"
            )

        except PermissionError:
            logger.error("PDF permission denied")
            QMessageBox.critical(
                self.ui,
                "Export Failed",
                "Cannot overwrite the PDF file.\n\n"
                "Please close the PDF if it is already open."
            )

        except Exception as e:
            logger.error(f"PDF export failed: {e}")
            QMessageBox.critical(
                self.ui,
                "Export Failed",
                str(e)
            )

